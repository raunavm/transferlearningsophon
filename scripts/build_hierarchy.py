#!/usr/bin/env python3
"""Build the `hierarchy/` artefact set for the label-granularity study.

Deterministic and re-runnable: every value is derived from the pinned Sophon
data config, or from an explicitly labelled construction decision recorded in
the output.  Re-running on the same inputs reproduces byte-identical CSVs.

Pinned inputs
-------------
  sophon @ 9dd6dd6a261aa6d5fd2e56f015068127b36854f9
  data/JetClassII/JetClassII_full.yaml
    line 129      -> the authoritative 188-class name list (comment block)
    lines 28-61   -> the 30 reweighting-group index expressions
    lines 151-152 -> reweight_vars bin edges
    lines 154-161 -> reweight_classes (order)
    lines 163-170 -> class_weights

Usage
-----
  python3 scripts/build_hierarchy.py [--sophon DIR] [--out DIR]
"""

from __future__ import annotations

import argparse
import csv
import datetime
import hashlib
import io
import math
import os
import platform
import re
import subprocess
import sys
import zipfile
from collections import OrderedDict
from fractions import Fraction

# --------------------------------------------------------------------------
# constants: pinned expectations that the script asserts against the file
# --------------------------------------------------------------------------

SOPHON_COMMIT = "9dd6dd6a261aa6d5fd2e56f015068127b36854f9"
YAML_REL = "data/JetClassII/JetClassII_full.yaml"

LINE_CLASS_NAMES = 129          # 1-based
LINE_REWEIGHT_FIRST = 28        # 1-based, label_X_QQ
LINE_REWEIGHT_LAST = 61         # 1-based, label_QCD
N_CLASSES = 188
BLOCK_2P = (0, 15)              # [lo, hi)
BLOCK_34P = (15, 161)
BLOCK_QCD = (161, 188)
CLASS_WEIGHTS_SUM = 1.73175
N_REWEIGHT_GROUPS = 30
N_PT_BINS = 14
N_SDMASS_BINS = 48

RAND42_SEEDS = (42, 43, 44)     # explicit, recorded in the outputs

ASSERTIONS: list[dict] = []     # every check, PASS/FAIL, becomes 04_*.csv


def check(name, category, passed, detail, chain_edge="", expectation="MUST_HOLD"):
    """`result` is the truth of the check. `expectation` says what the study
    requires. `verdict` is OK unless the two disagree -- so a human can filter
    on `verdict = DEFECT` and see only real problems, not the 1800 pairwise
    non-relations that are simply true answers to unrequired questions."""
    if expectation == "MUST_HOLD":
        verdict = "OK" if passed else "DEFECT"
    elif expectation == "MUST_NOT_HOLD":
        verdict = "OK (expected failure)" if not passed else "DEFECT"
    else:
        verdict = "informational"
    ASSERTIONS.append(
        dict(assertion=name, category=category, chain_edge=chain_edge,
             expectation=expectation, result="PASS" if passed else "FAIL",
             verdict=verdict, detail=detail)
    )
    return passed


def hard(name, category, passed, detail, chain_edge=""):
    """An assertion whose failure invalidates the artefact."""
    check(name, category, passed, detail, chain_edge)
    if not passed:
        raise AssertionError(f"{name}: {detail}")


# --------------------------------------------------------------------------
# 1. parse the pinned YAML (as text -- the class list lives in a comment)
# --------------------------------------------------------------------------

SAFE_EXPR = re.compile(r"^[\s()0-9&|<>=jet_labl]+$")


def parse_yaml(path):
    raw = open(path, encoding="utf-8").read()
    lines = raw.split("\n")

    # --- 188 class names, from the comment block at line 129 ---------------
    line = lines[LINE_CLASS_NAMES - 1]
    hard("class_name_list_line_129", "provenance",
         "a full list of label names" in line,
         f"YAML line {LINE_CLASS_NAMES} is the class-name comment block")
    m = re.search(r"\[(.*)\]", line)
    hard("class_name_list_bracketed", "provenance", m is not None,
         "class-name list is a single closed [...] on one line")
    names = [t.strip() for t in m.group(1).split(",")]
    hard("class_name_count_188", "structure", len(names) == N_CLASSES,
         f"parsed {len(names)} class names from line {LINE_CLASS_NAMES}")
    hard("class_names_unique", "structure", len(set(names)) == N_CLASSES,
         "all 188 class names distinct")

    # --- 30 reweighting-group expressions, lines 28-61 --------------------
    rw = OrderedDict()
    rw_lines = {}
    pat = re.compile(r"^\s+(label_[A-Za-z_]+):\s*(\(.*\))\s*$")
    for ln in range(LINE_REWEIGHT_FIRST, LINE_REWEIGHT_LAST + 1):
        mm = pat.match(lines[ln - 1])
        if mm:
            rw[mm.group(1)] = mm.group(2)
            rw_lines[mm.group(1)] = ln
    hard("reweight_group_count_30", "structure", len(rw) == N_REWEIGHT_GROUPS,
         f"parsed {len(rw)} reweighting-group expressions from YAML lines "
         f"{LINE_REWEIGHT_FIRST}-{LINE_REWEIGHT_LAST}")

    # evaluate each expression over jet_label = 0..187
    rw_members = OrderedDict()
    for gname, expr in rw.items():
        hard(f"reweight_expr_safe[{gname}]", "provenance",
             bool(SAFE_EXPR.match(expr)),
             "expression contains only jet_label comparisons and & | ()")
        members = [i for i in range(N_CLASSES)
                   if eval(expr, {"__builtins__": {}}, {"jet_label": i})]
        rw_members[gname] = members

    # --- reweight_classes order and class_weights -------------------------
    def bracket_block(start_key, end_key):
        i0 = next(i for i, l in enumerate(lines) if l.strip().startswith(start_key))
        i1 = next(i for i, l in enumerate(lines) if l.strip().startswith(end_key))
        return "\n".join(lines[i0:i1]), i0 + 1, i1

    blk, _, _ = bracket_block("reweight_classes:", "class_weights:")
    rw_order = [t.strip() for t in re.search(r"\[(.*)\]", blk, re.S).group(1).split(",")
                if t.strip()]
    blk, _, _ = bracket_block("class_weights:", "reweight_hists:")
    weights_raw = [t.strip() for t in
                   re.search(r"\[(.*)\]", blk, re.S).group(1).split(",") if t.strip()]
    weights = [float(t) for t in weights_raw]

    hard("reweight_classes_count_30", "structure", len(rw_order) == N_REWEIGHT_GROUPS,
         f"reweight_classes lists {len(rw_order)} names")
    hard("class_weights_count_30", "structure", len(weights) == N_REWEIGHT_GROUPS,
         f"class_weights lists {len(weights)} values")
    hard("reweight_classes_match_new_variables", "structure",
         rw_order == list(rw.keys()),
         "reweight_classes order == new_variables definition order")
    wsum = sum(weights)
    hard("class_weights_sum_1.73175", "structure",
         abs(wsum - CLASS_WEIGHTS_SUM) < 1e-9,
         f"sum(class_weights) = {wsum:.10g} (expected {CLASS_WEIGHTS_SUM})")

    # --- reweight_vars bin edges and selection ----------------------------
    def edges(varname):
        l = next(l for l in lines if l.strip().startswith(varname + ":"))
        return [float(t) for t in re.search(r"\[(.*)\]", l).group(1).split(",")]

    pt_edges = edges("jet_pt")
    sd_edges = edges("jet_sdmass")
    hard("reweight_pt_bins_14", "structure", len(pt_edges) - 1 == N_PT_BINS,
         f"jet_pt: {len(pt_edges)} edges -> {len(pt_edges)-1} bins over "
         f"{pt_edges[0]:g}-{pt_edges[-1]:g}")
    hard("reweight_sdmass_bins_48", "structure", len(sd_edges) - 1 == N_SDMASS_BINS,
         f"jet_sdmass: {len(sd_edges)} edges -> {len(sd_edges)-1} bins over "
         f"{sd_edges[0]:g}-{sd_edges[-1]:g}")
    sel = lines[3].strip()
    hard("selection_line", "provenance",
         sel == "(jet_pt > 200) & (jet_pt < 2500) & (jet_sdmass > 20) & (jet_sdmass < 500)",
         f"selection = {sel}")

    return dict(names=names, rw_members=rw_members, rw_lines=rw_lines,
                weights=weights, weights_raw=weights_raw, rw_order=rw_order,
                pt_edges=pt_edges, sd_edges=sd_edges, selection=sel, raw=raw)


# --------------------------------------------------------------------------
# 2. per-class physical attributes
# --------------------------------------------------------------------------

TOKENS = ["light", "tauh", "taue", "taum", "tau", "b", "c", "s", "q", "g", "e", "m", "v"]

TOKEN_LABEL = {
    "b": "b quark", "c": "c quark", "s": "s quark", "q": "light quark (u/d)",
    "g": "gluon", "e": "electron", "m": "muon", "v": "neutrino",
    "tauh": "hadronic tau", "taue": "tau->e", "taum": "tau->mu",
    "tau": "tau", "light": "light partons only",
}


def tokenize(suffix):
    toks, i = [], 0
    while i < len(suffix):
        for t in TOKENS:
            if suffix.startswith(t, i):
                toks.append(t)
                i += len(t)
                break
        else:
            raise ValueError(f"cannot tokenize {suffix!r} at position {i}")
    return toks


def class_attrs(idx, name, rw_group_name):
    if name.startswith("label_X_YY_"):
        block, suffix = "res34p", name[len("label_X_YY_"):]
    elif name.startswith("label_X_"):
        block, suffix = "res2p", name[len("label_X_"):]
    elif name.startswith("label_QCD_"):
        block, suffix = "qcd", name[len("label_QCD_"):]
    else:
        raise ValueError(name)

    toks = tokenize(suffix)
    has_b = "b" in toks
    has_c = "c" in toks
    has_s = "s" in toks
    has_g = "g" in toks
    has_tau = any(t.startswith("tau") for t in toks)
    has_nu = "v" in toks
    has_lepton = has_tau or "e" in toks or "m" in toks

    if block == "qcd":
        n_obj = n_vis = ""
        topo = "QCD"
    else:
        n_obj = len(toks)
        n_vis = len(toks) - toks.count("v")
        topo = rw_group_name.replace("label_X_YY_", "").replace("label_X_", "")

    if has_b:
        lhf = "b"
    elif has_c:
        lhf = "c"
    elif has_s:
        lhf = "s"
    elif has_g or "q" in toks or suffix == "light":
        lhf = "light"
    else:
        lhf = "none"

    return dict(
        jet_label=idx, class_name=name, block=block, topology_code=topo,
        tokens="+".join(toks), n_objects=n_obj, n_visible_objects=n_vis,
        has_b=int(has_b), has_c=int(has_c), has_s=int(has_s), has_g=int(has_g),
        has_lepton=int(has_lepton), has_tau=int(has_tau), has_nu=int(has_nu),
        leading_heavy_flavour=lhf,
        bs_relevant=int(has_b and has_s),
    )


# --------------------------------------------------------------------------
# 3. rung construction
# --------------------------------------------------------------------------

# ---- R = 9, two-prong: v3's 3 / 3 / 4 / 5 --------------------------------
R9_2P = [
    ("2P_QQ_b", [0, 4, 6],
     "Two-prong X->QQ' containing a b: b-hadron decay length gives displaced "
     "vertices and high track multiplicity, the single strongest flavour handle "
     "inside a large-R jet."),
    ("2P_QQ_c", [1, 5, 7],
     "Two-prong X->QQ' with c but no b: intermediate decay length and impact "
     "parameter, separable from both b and light."),
    ("2P_QQ_light_gluon", [2, 3, 8, 9],
     "Two-prong X->QQ'/gg with no b or c: prompt light-quark and gluon prongs, "
     "distinguished only by radiation pattern, so merged."),
    ("2P_leptonic", [10, 11, 12, 13, 14],
     "Two-prong X->leptons (ee, mm) or X->tautau: lepton/tau prongs replace "
     "hadronic prongs; lepton flavour and tau decay mode contracted "
     "(contraction step ii)."),
]

# ---- R = 9, 3/4-prong: v3's 27 / 25 / 29 / 35 / 30 ----------------------
R9_34P_MAP = OrderedDict([
    ("34P_4prong_hadronic", (
        ["QQQQ", "QQgg", "gggg"],
        "X->YY with four visible hadronic prongs; all quark flavour and all "
        "quark/gluon distinctions contracted (steps i-ii complete), prong count "
        "and hadronic character preserved.")),
    ("34P_4prong_leptonic", (
        ["QQll", "QQtauhtaul", "QQtauhtauh", "ggll", "ggtauhtaul", "ggtauhtauh"],
        "X->YY with two hadronic prongs plus two charged leptons/taus, four "
        "visible objects; lepton flavour and tau decay mode contracted.")),
    ("34P_3prong_hadronic", (
        ["QQQ", "QQg", "Qgg", "ggg"],
        "X->YY with three visible hadronic prongs (one Y decays to a single "
        "parton); quark flavour and quark/gluon contracted.")),
    ("34P_3prong_leptonic", (
        ["QQl", "Qll", "ggl", "gll", "Qtauhtaul", "Qtauhtauh",
         "gtauhtaul", "gtauhtauh"],
        "X->YY with three visible objects of which at least one is a charged "
        "lepton or tau; lepton flavour and tau decay mode contracted.")),
    ("34P_semileptonic_nu", (
        ["QQlv", "QQtaulv", "QQtauhv"],
        "X->YY with a leptonically decaying W-like Y (l-nu or tau-nu): the "
        "neutrino is invisible, so three of four objects are reconstructed and "
        "the jet carries missing momentum. Kept separate from the fully visible "
        "3-prong block because the invisible prong is a distinct signature.")),
])

# ---- R = 15 -------------------------------------------------------------
R15_2P = [
    ("2P_QQ_b", [0, 4, 6], R9_2P[0][2]),
    ("2P_QQ_c", [1, 5, 7], R9_2P[1][2]),
    ("2P_QQ_light_gluon", [2, 3, 8, 9], R9_2P[2][2]),
    ("2P_ll", [10, 11],
     "Two-prong X->ee / X->mumu: two prompt light leptons; e/mu flavour "
     "contracted, tau final states held separate at this rung."),
    ("2P_tautau", [12, 13, 14],
     "Two-prong X->tautau (any tau decay mode): tau prongs carry displaced, "
     "low-multiplicity substructure unlike prompt e/mu, so separated from 2P_ll "
     "until R=9."),
]

R15_34P_MAP = OrderedDict([
    ("34P_4prong_had_quarks", (["QQQQ"],
     "X->YY, four quark prongs, no gluon: four resolvable hadronic subjets.")),
    ("34P_4prong_had_gluon", (["QQgg", "gggg"],
     "X->YY, four hadronic prongs with at least one gluon-initiated Y: "
     "gluon prongs have wider, higher-multiplicity radiation than quark prongs.")),
    ("34P_4prong_lightlep", (["QQll", "ggll"],
     "X->YY, two hadronic prongs plus two prompt light leptons (e/mu).")),
    ("34P_4prong_tau", (["QQtauhtaul", "QQtauhtauh", "ggtauhtaul", "ggtauhtauh"],
     "X->YY, two hadronic prongs plus a tau pair; tau substructure held "
     "separate from prompt e/mu until R=9.")),
    ("34P_3prong_had_quarks", (["QQQ"],
     "X->YY, three quark prongs, no gluon.")),
    ("34P_3prong_had_gluon", (["QQg", "Qgg", "ggg"],
     "X->YY, three hadronic prongs with at least one gluon.")),
    ("34P_3prong_lightlep", (["QQl", "Qll", "ggl", "gll"],
     "X->YY, three visible objects including at least one prompt light lepton.")),
    ("34P_3prong_tau", (["Qtauhtaul", "Qtauhtauh", "gtauhtaul", "gtauhtauh"],
     "X->YY, three visible objects including a tau pair.")),
    ("34P_semilep_lnu", (["QQlv"],
     "X->YY with Y->l-nu, l = e/mu: prompt lepton plus missing momentum.")),
    ("34P_semilep_taunu", (["QQtaulv", "QQtauhv"],
     "X->YY with Y->tau-nu: tau substructure plus missing momentum, held "
     "separate from the prompt-lepton case until R=9.")),
])

# ---- R = 45, two-prong -------------------------------------------------
R45_2P = [
    ("2P_QQ_b", [0, 4, 6], R9_2P[0][2]),
    ("2P_QQ_c", [1, 5, 7], R9_2P[1][2]),
    ("2P_QQ_light", [2, 3, 8],
     "Two-prong X->QQ' with no b or c (ss, qq, sq): prompt light-quark prongs, "
     "separated from gg at this rung."),
    ("2P_gg", [9], "Two-prong X->gg: gluon prongs, wider and higher "
     "multiplicity than light-quark prongs."),
    ("2P_ll", [10, 11], "Two-prong X->ee / X->mumu; e/mu flavour contracted "
     "(this is the reweighting group label_X_ll)."),
    ("2P_tauhtaul", [12, 13], "Two-prong X->tautau with one hadronic and one "
     "leptonic tau (reweighting group label_X_tauhtaul)."),
    ("2P_tauhtauh", [14], "Two-prong X->tautau, both taus hadronic "
     "(reweighting group label_X_tauhtauh)."),
]


def build_partition(spec_groups, universe, label):
    """spec_groups: list of (name, members, rationale). Validate & return."""
    seen = set()
    for nm, mem, _ in spec_groups:
        for i in mem:
            hard(f"{label}_no_double_assignment", "structure", i not in seen,
                 f"class {i} assigned twice in {label}")
            seen.add(i)
    hard(f"{label}_covers_universe", "structure", seen == set(universe),
         f"{label} covers exactly {len(universe)} indices "
         f"(missing {sorted(set(universe) - seen)}, extra {sorted(seen - set(universe))})")
    # deterministic id order: by smallest member index
    ordered = sorted(spec_groups, key=lambda g: min(g[1]))
    return [dict(group_id=k, group_name=nm, members=sorted(mem), rationale=r)
            for k, (nm, mem, r) in enumerate(ordered)]


def topo_groups(topo_map, topo_to_indices, prefix_note):
    out = []
    for gname, (topos, rat) in topo_map.items():
        mem = []
        for t in topos:
            hard(f"topology_known[{t}]", "structure", t in topo_to_indices,
                 f"topology code {t} exists in the reweighting groups")
            mem.extend(topo_to_indices[t])
        out.append((gname, mem, rat + " " + prefix_note))
    return out


def hash_order(items, seed, tag):
    """Deterministic permutation, stable across Python versions/platforms."""
    def key(i):
        return hashlib.sha256(f"{tag}|seed={seed}|idx={i}".encode()).hexdigest()
    return sorted(items, key=key)


# --------------------------------------------------------------------------
# 4. refinement algebra
# --------------------------------------------------------------------------

def as_map(groups):
    m = {}
    for g in groups:
        for i in g["members"]:
            m[i] = g["group_id"]
    return m


def refines(fine, coarse):
    """True if `coarse` is a coarsening of `fine`: every fine group lies inside
    exactly one coarse group.  Returns (bool, detail)."""
    cm = as_map(coarse)
    bad = []
    for g in fine:
        tgt = {cm[i] for i in g["members"]}
        if len(tgt) != 1:
            bad.append(g["group_name"])
    if bad:
        return False, (f"{len(bad)} group(s) of the finer partition straddle a "
                       f"coarser boundary: {', '.join(bad[:6])}"
                       + (" ..." if len(bad) > 6 else ""))
    return True, (f"all {len(fine)} groups of the finer partition lie inside "
                  f"exactly one group of the coarser partition")


def compact_ranges(idxs):
    idxs = sorted(idxs)
    out, i = [], 0
    while i < len(idxs):
        j = i
        while j + 1 < len(idxs) and idxs[j + 1] == idxs[j] + 1:
            j += 1
        out.append(str(idxs[i]) if i == j else f"{idxs[i]}-{idxs[j]}")
        i = j + 1
    return ",".join(out)


# --------------------------------------------------------------------------
# 5. main
# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sophon", default=os.path.expanduser("~/dev/sophon"))
    ap.add_argument("--out", default=os.path.join(os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))), "hierarchy"))
    ap.add_argument("--handoff", default=os.path.join(os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))), "handoff"))
    args = ap.parse_args()

    repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    yaml_path = os.path.join(args.sophon, YAML_REL)
    hard("sophon_yaml_present", "provenance", os.path.exists(yaml_path),
         f"{yaml_path}")

    head = subprocess.run(["git", "-C", args.sophon, "rev-parse", "HEAD"],
                          capture_output=True, text=True).stdout.strip()
    hard("sophon_commit_pinned", "provenance", head == SOPHON_COMMIT,
         f"HEAD of {args.sophon} = {head} (expected {SOPHON_COMMIT})")
    dirty = subprocess.run(["git", "-C", args.sophon, "status", "--porcelain"],
                           capture_output=True, text=True).stdout.strip()
    tracked = [l for l in dirty.split("\n") if l and not l.startswith("??")]
    check("sophon_no_tracked_modifications", "provenance", not tracked,
          "no tracked file in the pinned clone is modified"
          + (f" (untracked cruft present but irrelevant: "
             f"{len(dirty.splitlines())} file(s), all ??)" if dirty else "")
          if not tracked else f"TRACKED MODIFICATIONS: {tracked}")

    # The config must be byte-identical across every clone, or "the pin" is a
    # fiction.
    yaml_hashes = {}
    for clone in ("~/sophon", "~/dev/sophon", "~/e1_refs/sophon"):
        p = os.path.join(os.path.expanduser(clone), YAML_REL)
        if os.path.exists(p):
            yaml_hashes[clone] = sha256(p)
    hard("sophon_yaml_identical_across_clones", "provenance",
         len(set(yaml_hashes.values())) == 1,
         f"{YAML_REL} is byte-identical across {len(yaml_hashes)} clones "
         f"({', '.join(yaml_hashes)}), sha256 "
         f"{next(iter(yaml_hashes.values()))}")

    Y = parse_yaml(yaml_path)
    names = Y["names"]
    rw_members = Y["rw_members"]
    weights = dict(zip(Y["rw_order"], Y["weights"]))
    W = sum(Y["weights"])

    # ---- reweighting-group bookkeeping ---------------------------------
    rw_of = {}
    for g, mem in rw_members.items():
        for i in mem:
            hard("reweight_groups_disjoint", "structure", i not in rw_of,
                 f"class {i} in two reweighting groups")
            rw_of[i] = g
    hard("reweight_groups_cover_188", "structure",
         set(rw_of) == set(range(N_CLASSES)),
         f"the 30 reweighting groups partition all {N_CLASSES} classes")
    res_rw = [g for g in rw_members if g != "label_QCD"]
    hard("reweight_29_cover_resonant", "structure",
         sorted(i for g in res_rw for i in rw_members[g]) == list(range(161)),
         "the 29 label_X_* reweighting groups cover exactly indices 0-160")
    hard("reweight_qcd_cover", "structure",
         rw_members["label_QCD"] == list(range(161, 188)),
         "label_QCD covers exactly indices 161-187")

    # ---- block boundaries ---------------------------------------------
    hard("block_2p_names", "structure",
         all(names[i].startswith("label_X_") and not names[i].startswith("label_X_YY_")
             for i in range(*BLOCK_2P)),
         "indices 0-14 are all two-prong label_X_* classes")
    hard("block_34p_names", "structure",
         all(names[i].startswith("label_X_YY_") for i in range(*BLOCK_34P)),
         "indices 15-160 are all 3/4-prong label_X_YY_* classes")
    hard("block_qcd_names", "structure",
         all(names[i].startswith("label_QCD_") for i in range(*BLOCK_QCD)),
         "indices 161-187 are all label_QCD_* classes")

    # ---- QCD stream share vs node share -------------------------------
    qcd_stream = weights["label_QCD"] / W
    qcd_nodes = 27 / N_CLASSES
    hard("qcd_stream_share", "statistics", abs(qcd_stream - 0.144) < 0.001,
         f"QCD = {weights['label_QCD']}/{W:.5f} = {qcd_stream*100:.4f}% of the "
         f"sampled training stream")
    hard("qcd_node_share", "statistics", abs(qcd_nodes - 0.1436) < 0.0001,
         f"QCD = 27/188 = {qcd_nodes*100:.4f}% of vocabulary nodes")
    check("qcd_stream_exceeds_node_share", "statistics", qcd_stream > qcd_nodes,
          f"stream {qcd_stream*100:.4f}% > nodes {qcd_nodes*100:.4f}% "
          f"(difference {100*(qcd_stream-qcd_nodes):.4f} pp)")

    # ---- per-class attributes -----------------------------------------
    attrs = [class_attrs(i, names[i], rw_of[i]) for i in range(N_CLASSES)]

    # nominal (uniform-within-reweighting-group) and exact-upper-bound shares
    for a in attrs:
        g = rw_of[a["jet_label"]]
        n_g = len(rw_members[g])
        a["reweight_group_name"] = g
        a["reweight_group_n_classes"] = n_g
        a["reweight_group_weight"] = weights[g]
        a["reweight_group_stream_share"] = weights[g] / W
        a["class_stream_share_nominal"] = weights[g] / W / n_g
        a["class_stream_share_upper_bound"] = weights[g] / W
        # --- S5 rarity-slope x-variable, full precision -----------------
        a["class_weight_share_frac"] = weights[g] / W / n_g
        a["class_weight_share_pct"] = 100.0 * weights[g] / W / n_g
        a["log_class_weight_share"] = math.log(weights[g] / W / n_g)
        a["log_class_weight_share_upper_bound"] = math.log(weights[g] / W)

    nominal = {a["jet_label"]: a["class_stream_share_nominal"] for a in attrs}

    topo_to_indices = OrderedDict()
    for g, mem in rw_members.items():
        if g == "label_QCD":
            continue
        t = g.replace("label_X_YY_", "").replace("label_X_", "")
        topo_to_indices[t] = list(mem)

    # =================================================================
    #  R = 161 (identity)
    # =================================================================
    R161 = build_partition(
        [(names[i], [i], "Native JetClass-II class; identity partition, no "
          "contraction applied.") for i in range(161)],
        range(161), "R161")

    # =================================================================
    #  R = 29 (exactly the 29 label_X_* reweighting groups)
    # =================================================================
    R29 = build_partition(
        [(g.replace("label_X_YY_", "34P_").replace("label_X_", "2P_"),
          list(rw_members[g]),
          "Sophon reweighting group " + g + " (YAML line "
          f"{Y['rw_lines'][g]}): all quark-flavour distinctions within fixed "
          "prong count and topology are contracted (contraction step i, "
          "complete). This is the unique rung at which the flat (pT, mSD) "
          "sampling weights are class-wise flat.")
         for g in res_rw],
        range(161), "R29")

    # =================================================================
    #  R = 15
    # =================================================================
    R15 = build_partition(
        list(R15_2P) + topo_groups(
            R15_34P_MAP, topo_to_indices,
            "Quark flavour fully contracted; prong count, gluon content and "
            "tau-vs-prompt-lepton preserved."),
        range(161), "R15")

    # =================================================================
    #  R = 9  (v3's nine resonant groups)
    # =================================================================
    R9 = build_partition(
        list(R9_2P) + topo_groups(
            R9_34P_MAP, topo_to_indices, ""),
        range(161), "R9")

    hard("R9_two_prong_sizes_3_3_4_5", "plan-continuity",
         sorted(len(g["members"]) for g in R9
                if max(g["members"]) < 15) == [3, 3, 4, 5],
         "R=9 two-prong group sizes = "
         + str(sorted(len(g["members"]) for g in R9 if max(g["members"]) < 15))
         + " (PLAN.md line 105 requires 3, 3, 4, 5)")
    r9_34 = [len(g["members"]) for g in R9 if min(g["members"]) >= 15]
    hard("R9_34prong_sizes_27_25_29_35_30", "plan-continuity",
         sorted(r9_34) == sorted([27, 25, 29, 35, 30]),
         f"R=9 3/4-prong group sizes = {r9_34} "
         "(PLAN.md line 105 requires 27, 25, 29, 35, 30)")

    # =================================================================
    #  R = 45 -- refinement of R=29 by b-content, derived not hand-picked
    # =================================================================
    # Candidate splits: every reweighting group with both b and non-b members.
    splittable = []
    for g in res_rw:
        mem = rw_members[g]
        b = [i for i in mem if attrs[i]["has_b"]]
        nb = [i for i in mem if not attrs[i]["has_b"]]
        if b and nb:
            splittable.append((g, b, nb))
    # The two-prong QQ group must additionally resolve c-vs-light because R=9
    # (v3, frozen) separates them; that split is forced, not optional.
    forced_2p = {"label_X_QQ"}
    optional = [s for s in splittable if s[0] not in forced_2p]

    n_forced_2p_blocks = len(R45_2P)                       # 7
    n_34_base = len([g for g in res_rw if g not in forced_2p and
                     g not in ("label_X_gg", "label_X_ll", "label_X_tauhtaul",
                               "label_X_tauhtauh")])       # 24
    target_extra = 45 - n_forced_2p_blocks - n_34_base
    n_drop = len(optional) - target_extra
    hard("R45_drop_count_is_one", "decision", n_drop == 1,
         f"{len(optional)} optional b-splits available in the 3/4-prong block, "
         f"{target_extra} can be taken to reach exactly 45 groups, so exactly "
         f"{n_drop} must be dropped")

    # Stated criterion for which split to drop: the smallest nominal stream
    # share of the b-subset, i.e. the least-populated (least learnable) split.
    def b_subset_share(s):
        g, b, _ = s
        return weights[g] / W * len(b) / len(rw_members[g])
    optional_sorted = sorted(optional, key=b_subset_share)
    dropped = optional_sorted[:n_drop]
    dropped_names = [d[0] for d in dropped]
    hard("R45_drop_criterion_unique", "decision",
         b_subset_share(optional_sorted[0]) < b_subset_share(optional_sorted[1]),
         "the minimum-b-subset-stream-share criterion has a unique minimum: "
         f"{optional_sorted[0][0]} at {100*b_subset_share(optional_sorted[0]):.4f}% "
         f"vs next {optional_sorted[1][0]} at "
         f"{100*b_subset_share(optional_sorted[1]):.4f}%")

    taken = {s[0]: s for s in optional if s[0] not in dropped_names}
    R45_spec = list(R45_2P)
    for g in res_rw:
        if g in forced_2p or g in ("label_X_gg", "label_X_ll",
                                   "label_X_tauhtaul", "label_X_tauhtauh"):
            continue
        base = g.replace("label_X_YY_", "34P_")
        rat_tail = (" Reweighting group " + g + "; R=45 restores one level of "
                    "quark flavour, namely b-content.")
        if g in taken:
            _, b, nb = taken[g]
            R45_spec.append((base + "_b", b,
                             "At least one b quark among the hadronic prongs: "
                             "b-hadron decay length gives displaced secondary "
                             "vertices resolvable inside a large-R jet." + rat_tail))
            R45_spec.append((base + "_nob", nb,
                             "No b quark among the hadronic prongs; c/s/light "
                             "and gluon flavour still contracted." + rat_tail))
        else:
            why = (" b-content NOT resolved here: this group's b-subset carries "
                   f"the smallest nominal stream share of any splittable group "
                   f"({100*b_subset_share(taken.get(g) or [s for s in optional if s[0]==g][0]):.4f}% "
                   "of the training stream), so the split would be the least "
                   "populated in the vocabulary; dropping exactly this one split "
                   "is what makes the group count exactly 45."
                   ) if g in dropped_names else (
                   " b-content cannot be resolved here: the group contains no "
                   "b quark, or every member contains one.")
            R45_spec.append((base, list(rw_members[g]),
                             "Reweighting group " + g + " kept whole at R=45." + why))
    R45 = build_partition(R45_spec, range(161), "R45")
    hard("R45_group_count_45", "decision", len(R45) == 45,
         f"R=45 has {len(R45)} groups")

    # =================================================================
    #  QCD rungs
    # =================================================================
    Q27 = build_partition(
        [(names[i], [i],
          "Native JetClass-II QCD class, defined by the heavy-flavour partons "
          "(pT > 10 GeV) matched within dR < R0 of the jet axis; identity "
          "partition, indices 161-187 left unmerged.")
         for i in range(161, 188)], range(161, 188), "Q27")
    Q1 = build_partition(
        [("QCD_all", list(range(161, 188)),
          "All 27 QCD classes merged into one, exactly as Sophon's own "
          "reweighting group label_QCD does. Destroys every heavy-flavour "
          "distinction in the background.")], range(161, 188), "Q1")

    # =================================================================
    #  L187 -- label-holdout companion to endpoint S5.
    #  L188 with the two-prong `cs` class merged into `qq`.
    # =================================================================
    I_CS, I_QQ = names.index("label_X_cs"), names.index("label_X_qq")
    hard("L187_cs_index_is_5", "L187", I_CS == 5,
         f"label_X_cs is jet_label {I_CS}")
    hard("L187_qq_index_is_3", "L187", I_QQ == 3,
         f"label_X_qq is jet_label {I_QQ}")
    hard("L187_both_in_two_prong_block", "L187",
         I_CS < 15 and I_QQ < 15,
         f"both merged classes lie in the two-prong block 0-14 "
         f"(cs={I_CS}, qq={I_QQ})")
    hard("L187_both_in_same_reweight_group", "L187",
         rw_of[I_CS] == rw_of[I_QQ] == "label_X_QQ",
         f"label_X_cs and label_X_qq both sit inside reweighting group "
         f"label_X_QQ (jet_label 0-8, YAML line {Y['rw_lines']['label_X_QQ']}), "
         f"so the merge cannot split or span a reweighting group")

    L187_MERGE_RAT = (
        "Two-prong X->qq' with the cs final state merged in: the label-holdout "
        "arm for endpoint S5. Both members sit inside reweighting group "
        "label_X_QQ, so the merge changes the label map and nothing else.")
    L187res = build_partition(
        [(names[i], [i], "Native JetClass-II class, unchanged from L188.")
         for i in range(161) if i not in (I_CS, I_QQ)]
        + [("2P_qq_plus_cs", [I_QQ, I_CS], L187_MERGE_RAT)],
        range(161), "L187res")
    hard("L187_resonant_group_count_160", "L187", len(L187res) == 160,
         f"L187 has {len(L187res)} resonant groups (161 minus one merge)")
    L187full = build_partition(
        [(g["group_name"], g["members"], g["rationale"]) for g in L187res]
        + [(names[i], [i], "Native JetClass-II QCD class, unchanged; Q=27.")
           for i in range(161, 188)],
        range(188), "L187full")
    hard("L187_total_class_count_187", "L187", len(L187full) == 187,
         f"L187 vocabulary has {len(L187full)} classes (160 resonant + 27 QCD)")

    # The load-bearing claim: the sampling measure is EXACTLY unchanged.
    # Computed in exact rationals, so "identical" means identical and not
    # "identical to within float accumulation order".
    W_exact = sum(Fraction(t) for t in Y["weights_raw"])
    weights_exact = dict(zip(Y["rw_order"], (Fraction(t) for t in Y["weights_raw"])))
    nominal_exact = {i: weights_exact[rw_of[i]] / W_exact
                     / len(rw_members[rw_of[i]]) for i in range(N_CLASSES)}
    hard("class_weights_sum_exact", "structure",
         W_exact == Fraction("1.73175"),
         f"sum(class_weights) as an exact rational = {W_exact} "
         f"= {float(W_exact)!r}")

    def induced_rw_shares(partition, label):
        """Stream share the sampler assigns to each of the 30 reweighting
        groups, as induced by a vocabulary. Well defined only if every
        vocabulary class lies inside a single reweighting group."""
        out = OrderedDict((g, Fraction(0)) for g in rw_members)
        straddle = [g["group_name"] for g in partition
                    if len({rw_of[i] for i in g["members"]}) != 1]
        hard(f"{label}_every_class_within_one_reweight_group", "L187",
             not straddle,
             f"all {len(partition)} {label} vocabulary classes lie inside a "
             "single reweighting group, so the sampling measure per label is "
             "unambiguous"
             if not straddle else f"straddling classes: {straddle}")
        for grp in partition:
            out[rw_of[grp["members"][0]]] += sum(
                (nominal_exact[i] for i in grp["members"]), Fraction(0))
        return out

    L188full = build_partition(
        [(names[i], [i], "") for i in range(188)], range(188), "L188full")
    sh188 = induced_rw_shares(L188full, "L188")
    sh187 = induced_rw_shares(L187full, "L187")
    hard("L187_sampling_measure_exactly_unchanged", "L187",
         list(sh188.items()) == list(sh187.items()),
         "the realised 30-group stream-share vector is IDENTICAL under L188 and "
         "L187 -- exact equality on all 30 entries, not approximate -- because "
         "the cs->qq merge is contained inside label_X_QQ. L187+ therefore sees "
         "a byte-identical jet stream to L188+: the arm is a pure label-only "
         "intervention. Computed in exact rationals; max abs difference over "
         "the 30 groups = "
         + str(max(abs(sh188[g] - sh187[g]) for g in sh188)))
    changed = [g["group_name"] for g in L187full
               if sum((nominal_exact[i] for i in g["members"]), Fraction(0))
               != nominal_exact[g["members"][0]]]
    hard("L187_only_merged_node_share_changes", "L187",
         changed == ["2P_qq_plus_cs"],
         f"exactly one of the 187 vocabulary nodes has a stream share differing "
         f"from its L188 counterpart: {changed}. Its share is the exact sum of "
         f"the two merged classes ({nominal[I_QQ]:.9g} + {nominal[I_CS]:.9g} = "
         f"{nominal[I_QQ]+nominal[I_CS]:.9g}); the other 186 are untouched")

    # =================================================================
    #  L43 replication arm (resonant side)
    # =================================================================
    L43res = build_partition(
        [(names[i], [i], "Sophon Appendix B.2 two-prong class, kept verbatim.")
         for i in range(15)]
        + [("other_resonant", list(range(15, 161)),
            "All 146 3/4-prong resonant classes absorbed into a single "
            "'other resonant' class: Sophon's B.2 object has no 3/4-prong "
            "vocabulary, so the replication arm must dump them somewhere.")],
        range(161), "L43res")

    # =================================================================
    #  rand42 -- 3 draws, resonant block only, stratified by prong topology
    # =================================================================
    strata = OrderedDict([("res2p", list(range(0, 15))),
                          ("res34p", list(range(15, 161)))])
    r15_sizes = OrderedDict()
    for s, idxs in strata.items():
        lo, hi = min(idxs), max(idxs)
        r15_sizes[s] = [len(g["members"]) for g in R15
                        if lo <= min(g["members"]) <= hi]
    hard("rand42_size_template_matches_R15", "structure",
         sum(sum(v) for v in r15_sizes.values()) == 161,
         "rand42 block-size template taken from R=15 within each stratum: "
         + "; ".join(f"{k}={v}" for k, v in r15_sizes.items()))

    RAND = {}
    for d, seed in enumerate(RAND42_SEEDS, start=1):
        spec = []
        for s, idxs in strata.items():
            perm = hash_order(idxs, seed, f"rand42|draw{d}|{s}")
            pos = 0
            for k, size in enumerate(r15_sizes[s]):
                mem = perm[pos:pos + size]
                pos += size
                spec.append((f"rand_{s}_{k:02d}", mem,
                             f"RANDOM control group, draw {d} (seed {seed}), "
                             f"stratum {s}, size {size} copied from the R=15 "
                             "block-size distribution. No physical meaning by "
                             "construction: this is the count-versus-meaning "
                             "control."))
            hard(f"rand42_draw{d}_{s}_exhausted", "structure", pos == len(idxs),
                 f"draw {d} stratum {s}: {pos} of {len(idxs)} indices consumed")
        RAND[d] = build_partition(spec, range(161), f"rand42_d{d}")
        hard(f"rand42_draw{d}_group_count_15", "structure", len(RAND[d]) == 15,
             f"draw {d} has {len(RAND[d])} resonant groups")

    # =================================================================
    #  refinement assertion matrix
    # =================================================================
    res_parts = OrderedDict([("R161", R161), ("L187res", L187res), ("R45", R45),
                             ("R29", R29), ("R15", R15), ("R9", R9),
                             ("L43res", L43res),
                             ("rand42_d1", RAND[1]), ("rand42_d2", RAND[2]),
                             ("rand42_d3", RAND[3])])
    qcd_parts = OrderedDict([("Q27", Q27), ("Q1", Q1)])

    # Edges of the DELIVERED hierarchy (a DAG, not a total order) ...
    DELIVERED = {("R161", "R45"), ("R45", "R29"), ("R45", "R15"), ("R15", "R9")}
    # ... and its transitive closure, all of which must hold.
    REQUIRED = DELIVERED | {("R161", "R29"), ("R161", "R15"), ("R161", "R9"),
                            ("R45", "R9"),
                            # L187 is L188 with one merge, so it must sit
                            # directly under R161 and above R29.
                            ("R161", "L187res"), ("L187res", "R29")}
    # Edges the brief's total order 161>45>29>15>9 asks for that are provably
    # impossible: R=29 holds label_X_QQ (9 classes) whole, R=9's largest
    # two-prong group has 5 members.
    IMPOSSIBLE = {("R29", "R15"), ("R29", "R9")}

    for fname, fine in res_parts.items():
        for cname, coarse in res_parts.items():
            ok, det = refines(fine, coarse)
            pair = (fname, cname)
            if fname == cname:
                edge, exp = "identity", "MUST_HOLD"
                det = "identity: a partition is trivially a coarsening of itself"
            elif pair in DELIVERED:
                edge, exp = "DELIVERED CHAIN EDGE", "MUST_HOLD"
            elif pair in {("R161", "L187res"), ("L187res", "R29")}:
                edge, exp = "L187 holdout arm edge", "MUST_HOLD"
            elif pair in REQUIRED:
                edge, exp = "implied by delivered chain (transitive)", "MUST_HOLD"
            elif pair in IMPOSSIBLE:
                edge, exp = ("brief-requested edge, PROVABLY IMPOSSIBLE",
                             "MUST_NOT_HOLD")
            else:
                edge, exp = "", "not_required"
            check(f"refines[{fname} > {cname}]", "refinement", ok, det, edge, exp)

    for fname, fine in qcd_parts.items():
        for cname, coarse in qcd_parts.items():
            ok, det = refines(fine, coarse)
            if fname == cname:
                edge, exp = "identity", "MUST_HOLD"
                det = "identity"
            elif (fname, cname) == ("Q27", "Q1"):
                edge, exp = "DELIVERED CHAIN EDGE", "MUST_HOLD"
            else:
                edge, exp = "", "not_required"
            check(f"refines[{fname} > {cname}]", "refinement", ok, det, edge, exp)

    # Q=27 leaves 161-187 unmerged, for every Q=27 vocabulary
    for vname in ("L188", "L72", "L56", "L42", "L36", "L43", "rand42_d1",
                  "rand42_d2", "rand42_d3"):
        check(f"qcd_unmerged_at_Q27[{vname}]", "refinement",
              all(len(g["members"]) == 1 for g in Q27) and len(Q27) == 27,
              "all 27 QCD classes are singleton groups (indices 161-187 unmerged)")
    check("qcd_merged_at_Q1", "refinement",
          len(Q1) == 1 and len(Q1[0]["members"]) == 27,
          "at Q=1 all 27 QCD classes collapse into a single group")

    # =================================================================
    #  vocabularies and statistics
    # =================================================================
    def group_shares(groups):
        out = []
        for g in groups:
            nom = sum(nominal[i] for i in g["members"])
            touched = {rw_of[i] for i in g["members"]}
            full = [t for t in touched
                    if set(rw_members[t]).issubset(set(g["members"]))]
            lo = sum(weights[t] / W for t in full)
            hi = sum(weights[t] / W for t in touched)
            out.append((nom, lo, hi, sorted(touched)))
        return out

    VOCAB = OrderedDict()
    for rname, rpart in [("R161", R161), ("R45", R45), ("R29", R29),
                         ("R15", R15), ("R9", R9)]:
        for qname, qpart in [("Q27", Q27), ("Q1", Q1)]:
            label = {("R161", "Q27"): "L188", ("R45", "Q27"): "L72",
                     ("R29", "Q27"): "L56", ("R15", "Q27"): "L42",
                     ("R9", "Q27"): "L36", ("R161", "Q1"): "L162",
                     ("R45", "Q1"): "L46", ("R29", "Q1"): "L30",
                     ("R15", "Q1"): "L16", ("R9", "Q1"): "L10"}[(rname, qname)]
            VOCAB[label] = (rname, rpart, qname, qpart)
    VOCAB["L187"] = ("L187res", L187res, "Q27", Q27)
    VOCAB["L43"] = ("L43res", L43res, "Q27", Q27)
    for d in (1, 2, 3):
        VOCAB[f"rand42_d{d}"] = (f"rand42_d{d}", RAND[d], "Q27", Q27)

    H30 = -sum((w / W) * math.log(w / W) for w in Y["weights"])

    summary = []
    for label, (rname, rpart, qname, qpart) in VOCAB.items():
        gs = group_shares(rpart) + group_shares(qpart)
        shares = [s[0] for s in gs]
        n_exact = sum(1 for _, lo, hi, _ in gs if lo == hi)
        tot = sum(shares)
        p = [s / tot for s in shares]
        H = -sum(x * math.log(x) for x in p if x > 0)
        summary.append(dict(
            vocabulary=label, R=len(rpart), Q=len(qpart),
            N_classes=len(rpart) + len(qpart),
            resonant_rung=rname, qcd_rung=qname,
            entropy_nats=H, effective_class_count=math.exp(H),
            min_group_stream_share=min(shares), max_group_stream_share=max(shares),
            max_over_min_ratio=max(shares) / min(shares),
            ln_R=math.log(len(rpart)), R_linear=len(rpart),
            ln_N=math.log(len(rpart) + len(qpart)),
            entropy_x=H,
            n_groups_with_exact_share=n_exact,
            n_groups_share_bounded_only=len(gs) - n_exact,
            group_size_vector=",".join(str(len(g["members"])) for g in rpart),
        ))

    # ---- trend-contrast SE factors ------------------------------------
    def se_rows():
        rows = []
        designs = OrderedDict([
            ("core_ladder_Q27  (R=161,45,15,9)", ["L188", "L72", "L42", "L36"]),
            ("extended_Q27     (R=161,45,29,15,9)",
             ["L188", "L72", "L56", "L42", "L36"]),
        ])
        by_label = {s["vocabulary"]: s for s in summary}
        for dname, labels in designs.items():
            n = len(labels)
            xsets = OrderedDict([
                ("ln_R", [by_label[l]["ln_R"] for l in labels]),
                ("R_linear", [by_label[l]["R_linear"] for l in labels]),
                ("ln_N", [by_label[l]["ln_N"] for l in labels]),
                ("effective_entropy_nats", [by_label[l]["entropy_x"] for l in labels]),
                ("equal_spacing_rank", list(range(1, n + 1))),
            ])
            for xn, xs in xsets.items():
                xb = sum(xs) / n
                Sxx = sum((x - xb) ** 2 for x in xs)
                rng = max(xs) - min(xs)
                xstd = [(x - min(xs)) / rng for x in xs]
                xsb = sum(xstd) / n
                Sxx_std = sum((x - xsb) ** 2 for x in xstd)
                lev = [(x - xb) ** 2 / Sxx for x in xs]
                rows.append(dict(
                    design=dname, n_rungs=n, x_variable=xn,
                    rungs=" | ".join(labels),
                    x_values=" | ".join(f"{x:.6g}" for x in xs),
                    x_mean=xb, Sxx=Sxx,
                    se_factor_raw=1 / math.sqrt(Sxx),
                    Sxx_range_standardised=Sxx_std,
                    se_factor_range_standardised=1 / math.sqrt(Sxx_std),
                    max_leverage_fraction=max(lev),
                    max_leverage_rung=labels[lev.index(max(lev))],
                    min_leverage_fraction=min(lev),
                    leverage_by_rung=" | ".join(f"{v:.3f}" for v in lev),
                ))
        # rank within each design by standardised SE factor (lower = more power)
        for dname in designs:
            sub = [r for r in rows if r["design"] == dname]
            for k, r in enumerate(sorted(
                    sub, key=lambda r: r["se_factor_range_standardised"]), 1):
                r["power_rank_within_design"] = k
            for k, r in enumerate(sorted(
                    sub, key=lambda r: r["max_leverage_fraction"]), 1):
                r["balance_rank_within_design"] = k
        return rows

    SE = se_rows()

    # =================================================================
    #  b-and-s mechanism
    # =================================================================
    bs_classes = [a["jet_label"] for a in attrs if a["bs_relevant"]]
    check("bs_classes_found", "bs-mechanism", len(bs_classes) >= 3,
          f"{len(bs_classes)} classes carry both b and s content: "
          + ", ".join(names[i] for i in bs_classes))
    check("bs_absent_in_two_prong", "bs-mechanism",
          not any(i < 15 for i in bs_classes),
          "there is no two-prong label_X_bs class in the vocabulary; b+s "
          "content appears only in 3-prong resonant, 4-prong resonant and QCD "
          "classes, so 'genuinely unseen at every rung' is FALSE")

    # ---- correction-queue item 8a: "no new simulation" substitutions ----
    ITEM_8A = [("8a.1", "label_QCD_bs", 177),
               ("8a.2", "label_X_YY_qqbs", 124),
               ("8a.3", "label_X_YY_ccbs", 119)]
    for clause, cname, idx in ITEM_8A:
        present = cname in names
        at_idx = present and names.index(cname) == idx
        flagged = present and attrs[names.index(cname)]["bs_relevant"] == 1
        check(f"item_8a[{clause}] {cname} in released schema", "item-8a",
              present and at_idx and flagged,
              f"{cname} present in the pinned 188-class schema at jet_label "
              f"{names.index(cname) if present else 'MISSING'} "
              f"(expected {idx}), bs_relevant={1 if flagged else 0}")
    # 8a.4: X->bb projection needs nothing outside JetClass-II.
    bb_ok = names[0] == "label_X_bb"
    qcd_ok = all(names[i].startswith("label_QCD_") for i in range(161, 188))
    check("item_8a[8a.4] X->bb projection constructible in-corpus", "item-8a",
          bb_ok and qcd_ok,
          f"label_X_bb is jet_label 0 ({bb_ok}); a QCD-only background is "
          f"available from the same 188-class schema, jet_label 161-187, all 27 "
          f"present ({qcd_ok}). Signal and background therefore both live inside "
          f"JetClass-II and the projection is constructible in principle with no "
          f"external sample. NOTE: existence only -- the sealed final-eval split "
          f"was NOT read, and no occupancy was counted (blocker B4)")

    # =================================================================
    #  write outputs
    # =================================================================
    os.makedirs(args.out, exist_ok=True)
    os.makedirs(args.handoff, exist_ok=True)
    written = []

    def wcsv(fname, header, rows):
        path = os.path.join(args.out, fname)
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh, lineterminator="\n")
            w.writerow(header)
            for r in rows:
                w.writerow([r.get(h, "") if isinstance(r, dict) else r[header.index(h)]
                            for h in header])
        written.append(fname)
        return header, rows

    TABS = OrderedDict()

    # ---- 01_class_master.csv -----------------------------------------
    rung_cols = OrderedDict([("R161", R161), ("L187", L187full), ("R45", R45),
                             ("R29", R29), ("R15", R15), ("R9", R9),
                             ("Q27", Q27), ("Q1", Q1),
                             ("L43res", L43res), ("rand42_d1", RAND[1]),
                             ("rand42_d2", RAND[2]), ("rand42_d3", RAND[3])])
    maps = {k: (as_map(v), {g["group_id"]: g["group_name"] for g in v})
            for k, v in rung_cols.items()}

    master_header = [
        "jet_label", "class_name", "block", "topology_code", "tokens",
        "n_objects", "n_visible_objects",
        "has_b", "has_c", "has_s", "has_g", "has_lepton", "has_tau", "has_nu",
        "leading_heavy_flavour",
        "reweight_group_name", "reweight_group_n_classes",
        "reweight_group_weight", "reweight_group_stream_share",
        "class_stream_share_nominal", "class_stream_share_upper_bound",
        "class_weight_share_frac", "class_weight_share_pct",
        "log_class_weight_share", "log_class_weight_share_upper_bound",
    ]
    for k in rung_cols:
        master_header += [f"{k}_group_id", f"{k}_group_name"]
    master_header += ["bs_relevant", "physics_rationale"]

    master_rows = []
    for a in attrs:
        i = a["jet_label"]
        r = dict(a)
        for k in rung_cols:
            m, nm = maps[k]
            if i in m:
                r[f"{k}_group_id"] = m[i]
                r[f"{k}_group_name"] = nm[m[i]]
            else:
                r[f"{k}_group_id"] = ""
                r[f"{k}_group_name"] = "n/a (rung partitions the other block)"
        # one-line physics rationale
        objs = " + ".join(TOKEN_LABEL[t] for t in a["tokens"].split("+"))
        if a["block"] == "qcd":
            r["physics_rationale"] = (
                f"QCD multijet background; heavy-flavour partons matched to the "
                f"jet: {objs}. Distinguishable only by displaced vertices and "
                f"track kinematics; survives as its own node at Q=27 and is "
                f"destroyed at Q=1.")
        else:
            blk = "two-prong X->" if a["block"] == "res2p" else "X->YY with "
            coarse = maps["R9"][1][maps["R9"][0][i]]
            r["physics_rationale"] = (
                f"{blk}{objs}; topology {a['topology_code']}, "
                f"{a['n_visible_objects']} visible object(s), leading heavy "
                f"flavour {a['leading_heavy_flavour']}. Lands in {coarse} at the "
                f"coarsest resonant rung (R=9) because prong count and "
                f"hadronic-vs-leptonic character survive every contraction "
                f"applied above it.")
        master_rows.append(r)
    TABS["01_class_master"] = wcsv("01_class_master.csv", master_header, master_rows)

    # ---- 02_rung_<name>_groups.csv -----------------------------------
    CONTRACTION = {
        "R161": "none (identity)",
        "L187": "one merge only: two-prong label_X_cs into label_X_qq. Both sit "
                "inside reweighting group label_X_QQ, so the sampling measure is "
                "exactly unchanged and L187 vs L188 is a pure label-only "
                "intervention (label-holdout companion to endpoint S5)",
        "R45": "(i) quark flavour, partially: b-content resolved, c/s/light merged",
        "R29": "(i) quark flavour, complete",
        "R15": "(i) complete; (ii) lepton flavour and tau decay mode, partially "
               "(prompt e/mu merged, tau modes merged, prompt-vs-tau kept)",
        "R9": "(i) complete; (ii) complete; (iii) prong count NOT contracted; "
              "(iv) hadronic-vs-leptonic NOT contracted (two-prong block retains "
              "leading-heavy-flavour, which is why it is not a coarsening of R=29)",
        "Q27": "none (identity on indices 161-187)",
        "Q1": "all QCD heavy-flavour information contracted",
        "L187res": "see L187",
        "L43res": "two-prong identity; all 3/4-prong contracted to one node",
        "rand42_d1": "none physical: random control",
        "rand42_d2": "none physical: random control",
        "rand42_d3": "none physical: random control",
    }
    for k, groups in rung_cols.items():
        sh = group_shares(groups)
        hdr = ["group_id", "group_name", "n_classes", "member_jet_labels",
               "member_class_names", "contraction_applied_at_this_rung",
               "stream_share_nominal", "stream_share_exact_lower_bound",
               "stream_share_exact_upper_bound",
               "reweighting_groups_touched", "n_reweighting_groups_touched",
               "contains_b_and_s_class", "physics_rationale"]
        rows = []
        for g, (nom, lo, hi, touched) in zip(groups, sh):
            rows.append(dict(
                group_id=g["group_id"], group_name=g["group_name"],
                n_classes=len(g["members"]),
                member_jet_labels=compact_ranges(g["members"]),
                member_class_names=" ".join(names[i] for i in g["members"]),
                contraction_applied_at_this_rung=CONTRACTION[k],
                stream_share_nominal=nom,
                stream_share_exact_lower_bound=lo,
                stream_share_exact_upper_bound=hi,
                reweighting_groups_touched=" ".join(touched),
                n_reweighting_groups_touched=len(touched),
                contains_b_and_s_class=int(any(attrs[i]["bs_relevant"]
                                               for i in g["members"])),
                physics_rationale=g["rationale"].strip(),
            ))
        TABS[f"02_rung_{k}"] = wcsv(f"02_rung_{k}_groups.csv", hdr, rows)

    # ---- 03_rung_summary.csv -----------------------------------------
    hdr3 = ["vocabulary", "resonant_rung", "R", "qcd_rung", "Q", "N_classes",
            "entropy_nats", "effective_class_count",
            "min_group_stream_share", "max_group_stream_share",
            "max_over_min_ratio", "n_groups_with_exact_share",
            "n_groups_share_bounded_only", "ln_R", "R_linear", "ln_N",
            "entropy_x", "group_size_vector"]
    TABS["03_rung_summary"] = wcsv("03_rung_summary.csv", hdr3, summary)

    # ---- 03b_trend_se_factors.csv ------------------------------------
    hdr3b = ["design", "n_rungs", "x_variable", "rungs", "x_values", "x_mean",
             "Sxx", "se_factor_raw", "Sxx_range_standardised",
             "se_factor_range_standardised", "power_rank_within_design",
             "leverage_by_rung", "max_leverage_fraction", "max_leverage_rung",
             "min_leverage_fraction", "balance_rank_within_design"]
    TABS["03b_trend_se_factors"] = wcsv("03b_trend_se_factors.csv", hdr3b, SE)

    # ---- 04_refinement_assertions.csv --------------------------------
    hdr4 = ["assertion", "category", "chain_edge", "expectation", "result",
            "verdict", "detail"]
    TABS["04_refinement_assertions"] = wcsv("04_refinement_assertions.csv",
                                            hdr4, ASSERTIONS)

    # ---- 05_bs_mechanism.csv -----------------------------------------
    hdr5 = ["status", "jet_label", "class_name", "block", "topology_code",
            "tokens", "reweight_group_name"]
    for k in ("R161", "R45", "R29", "R15", "R9", "Q27", "Q1"):
        hdr5 += [f"{k}_group_id", f"{k}_group_name", f"{k}_group_n_classes",
                 f"{k}_isolated"]
    hdr5 += ["first_rung_where_bs_merges_away", "note"]

    rows5 = []
    for i in bs_classes:
        a = attrs[i]
        r = dict(status="present", jet_label=i, class_name=names[i],
                 block=a["block"], topology_code=a["topology_code"],
                 tokens=a["tokens"], reweight_group_name=a["reweight_group_name"])
        merged_at = ""
        for k in ("R161", "R45", "R29", "R15", "R9", "Q27", "Q1"):
            m, nm = maps[k]
            if i not in m:
                r[f"{k}_group_id"] = r[f"{k}_group_name"] = ""
                r[f"{k}_group_n_classes"] = r[f"{k}_isolated"] = ""
                continue
            gid = m[i]
            grp = next(g for g in rung_cols[k] if g["group_id"] == gid)
            iso = int(len(grp["members"]) == 1)
            r[f"{k}_group_id"] = gid
            r[f"{k}_group_name"] = nm[gid]
            r[f"{k}_group_n_classes"] = len(grp["members"])
            r[f"{k}_isolated"] = iso
            if not iso and merged_at == "":
                merged_at = k
        r["first_rung_where_bs_merges_away"] = merged_at
        r["note"] = ("b+s content in a "
                     + ("QCD jet" if a["block"] == "qcd" else
                        f"{a['n_visible_objects']}-prong resonance")
                     + "; separable from this rung upward only.")
        rows5.append(r)
    rows5.append(dict(
        status="absent_from_vocabulary", jet_label="", class_name="label_X_bs",
        block="res2p", topology_code="QQ", tokens="b+s",
        reweight_group_name="(would be label_X_QQ)",
        first_rung_where_bs_merges_away="",
        note="The two-prong b+s combination has NO class in JetClass-II: the "
             "two-prong QQ' list is bb, cc, ss, qq, bc, cs, bq, cq, sq -- bs is "
             "the one missing pair. MECHANISM: the neutral parent decays "
             "flavour-diagonally and the charged parent up-type-to-down-type, so "
             "b and s can only co-occur via two different parent decays, which "
             "requires at least three prongs. (Do NOT attribute this to an FCNC "
             "the generation lacks -- that wording is not in Sophon App. A, whose "
             "case (1) uses heft.) b+s content is accordingly present in the "
             "3-prong (bbs, bsq), 4-prong (bbss, ccbs, qqbs) and QCD (bbss, bbs, "
             "bss, bs) blocks, so b+s is NOT unseen at any rung."))
    TABS["05_bs_mechanism"] = wcsv("05_bs_mechanism.csv", hdr5, rows5)

    # ---- 00_README.md -------------------------------------------------
    readme = build_readme(args, Y, W, weights, qcd_stream, qcd_nodes, H30,
                          summary, SE, res_parts, dropped_names,
                          b_subset_share, optional, RAND42_SEEDS, r15_sizes,
                          bs_classes, names, nominal, I_QQ, I_CS)
    with open(os.path.join(args.out, "00_README.md"), "w", encoding="utf-8") as fh:
        fh.write(readme)
    written.insert(0, "00_README.md")

    # ---- xlsx workbook ------------------------------------------------
    xlsx_name = "hierarchy_artefact.xlsx"
    hash_rows = build_hash_rows(args, repo, yaml_path, head, written, xlsx_name)
    hdr6 = ["kind", "name", "path_or_ref", "value", "bytes",
            "byte_reproducible", "note"]
    build_xlsx(os.path.join(args.out, xlsx_name), TABS, hdr6, hash_rows)

    # ---- 06_hashes.csv (last: includes the workbook) -------------------
    xp = os.path.join(args.out, xlsx_name)
    hash_rows = hash_rows + [dict(
        kind="artefact", name=xlsx_name, path_or_ref=f"hierarchy/{xlsx_name}",
        value=sha256(xp), bytes=os.path.getsize(xp), byte_reproducible="yes",
        note="Excel rendering of the CSVs, one tab per file; zip member "
             "timestamps normalised to 1980-01-01 so the workbook is "
             "byte-reproducible for a fixed zlib version.")]
    wcsv("06_hashes.csv", hdr6, hash_rows)

    # ---- handoff/pins.md ---------------------------------------------
    with open(os.path.join(args.handoff, "pins.md"), "w", encoding="utf-8") as fh:
        fh.write(build_pins(args, repo, head))

    # ---- console report ----------------------------------------------
    defects = [a for a in ASSERTIONS if a["verdict"] == "DEFECT"]
    expfail = [a for a in ASSERTIONS
               if a["expectation"] == "MUST_NOT_HOLD" and a["result"] == "FAIL"]
    print(f"\nwrote {len(written)+1} files to {args.out}")
    for f in written + [xlsx_name]:
        p = os.path.join(args.out, f)
        print(f"  {sha256(p)}  {f}")
    print(f"\nassertions: {len(ASSERTIONS)} total, "
          f"{sum(1 for a in ASSERTIONS if a['expectation'] == 'MUST_HOLD')} "
          f"must-hold, {len(defects)} DEFECT, "
          f"{len(expfail)} expected-and-confirmed failures")
    for a in defects:
        print(f"  DEFECT {a['assertion']}: {a['detail']}")
    for a in expfail:
        print(f"  EXPECTED-FAIL {a['assertion']}: {a['detail']}")
    for nm, part in res_parts.items():
        print(f"  {nm:10s} sizes: "
              f"{[len(g['members']) for g in part]}")
    print()
    for s in summary:
        print(f"  {s['vocabulary']:10s} N={s['N_classes']:4d}  H={s['entropy_nats']:.5f} "
              f"nats  expH={s['effective_class_count']:8.3f}  "
              f"min={s['min_group_stream_share']:.3e} "
              f"max={s['max_group_stream_share']:.3e} "
              f"ratio={s['max_over_min_ratio']:.1f}")
    print(f"\n  H(30-group prior) = {H30:.6f} nats, exp(H) = {math.exp(H30):.4f}")
    print()
    for r in SE:
        print(f"  {r['design']}  {r['x_variable']:24s} "
              f"SE_raw={r['se_factor_raw']:.6g}  "
              f"SE_std={r['se_factor_range_standardised']:.6g}  "
              f"h_max={r['max_leverage_fraction']:.3f}@{r['max_leverage_rung']}  "
              f"rank={r['power_rank_within_design']}")
    return 0 if not defects else 1


# --------------------------------------------------------------------------
# helpers: hashing, workbook, README, pins
# --------------------------------------------------------------------------

def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def build_hash_rows(args, repo, yaml_path, head, written, xlsx_name):
    rows = []
    for f in written:
        p = os.path.join(args.out, f)
        rows.append(dict(kind="artefact", name=f, path_or_ref=f"hierarchy/{f}",
                         value=sha256(p), bytes=os.path.getsize(p),
                         byte_reproducible="yes", note=""))
    script = os.path.abspath(__file__)
    rows.append(dict(kind="generator", name="build_hierarchy.py",
                     path_or_ref="scripts/build_hierarchy.py",
                     value=sha256(script), bytes=os.path.getsize(script),
                     byte_reproducible="yes",
                     note="the only thing that produced these files"))
    rows.append(dict(kind="input", name="JetClassII_full.yaml",
                     path_or_ref=f"sophon@{head[:7]}:{YAML_REL}",
                     value=sha256(yaml_path), bytes=os.path.getsize(yaml_path),
                     byte_reproducible="yes",
                     note="every class name, reweighting group and class weight "
                          "is parsed from this file"))
    rows.append(dict(kind="commit", name="sophon", path_or_ref=args.sophon,
                     value=head, bytes="", byte_reproducible="n/a",
                     note="pinned analysis repo"))
    for clone in ("~/sophon", "~/dev/sophon", "~/e1_refs/sophon"):
        p = os.path.expanduser(clone)
        if os.path.isdir(p):
            h = subprocess.run(["git", "-C", p, "rev-parse", "HEAD"],
                               capture_output=True, text=True).stdout.strip()
            rows.append(dict(kind="commit", name=f"sophon clone {clone}",
                             path_or_ref=p, value=h, bytes="",
                             byte_reproducible="n/a",
                             note="agrees with pin" if h == head else "MISMATCH"))
    wc = os.path.expanduser("~/e1_refs/weaver-core")
    if os.path.isdir(wc):
        h = subprocess.run(["git", "-C", wc, "rev-parse", "HEAD"],
                           capture_output=True, text=True).stdout.strip()
        rows.append(dict(kind="commit", name="weaver-core (local clone)",
                         path_or_ref=wc, value=h, bytes="",
                         byte_reproducible="n/a",
                         note="branch dev/custom_train_eval; train_sophon.sh "
                              "does NOT pin a commit -- see handoff/pins.md"))
    ck = os.path.join(repo, "models/JetClassII_Sophon/model.pt")
    if os.path.exists(ck):
        rows.append(dict(kind="checkpoint", name="Sophon released model.pt",
                         path_or_ref="models/JetClassII_Sophon/model.pt",
                         value=sha256(ck), bytes=os.path.getsize(ck),
                         byte_reproducible="n/a",
                         note="downloaded from huggingface.co/jet-universe/"
                              "sophon/resolve/main/models/JetClassII_Sophon/"
                              "model.pt"))
    rows.append(dict(kind="note", name="self-exclusion",
                     path_or_ref="hierarchy/06_hashes.csv", value="", bytes="",
                     byte_reproducible="yes",
                     note="06_hashes.csv cannot contain its own hash. The "
                          "workbook's 06 tab is written before the workbook "
                          "exists and therefore lacks the workbook row that "
                          "this CSV has."))
    return rows


def build_xlsx(path, tabs, hdr6, hash_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    PALETTE = ["FFE8F1FF", "FFFFF2E0", "FFE9F7E9", "FFFDE9F3", "FFEFEAFB",
               "FFFFFBE0", "FFE4F5F7", "FFF6EBE2", "FFEDEDED", "FFE0F0E8",
               "FFFBE9E9", "FFEAF0FB"]
    wb = Workbook()
    wb.remove(wb.active)
    # Pin the document properties: openpyxl otherwise stamps wall-clock times
    # into docProps/core.xml, which would make the workbook non-reproducible.
    fixed = datetime.datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.creator = "scripts/build_hierarchy.py"
    wb.properties.lastModifiedBy = "scripts/build_hierarchy.py"
    allt = list(tabs.items()) + [("06_hashes", (hdr6, hash_rows))]
    for name, (hdr, rows) in allt:
        ws = wb.create_sheet(name[:31])
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FFD9D9D9")
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for r in rows:
            ws.append([r.get(h, "") if isinstance(r, dict)
                       else r[hdr.index(h)] for h in hdr])
        ws.freeze_panes = "A2"
        gid_cols = [k + 1 for k, h in enumerate(hdr)
                    if h.endswith("_group_id") or h == "group_id"]
        for ci in gid_cols:
            for ri in range(2, len(rows) + 2):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, int):
                    ws.cell(row=ri, column=ci).fill = PatternFill(
                        "solid", fgColor=PALETTE[v % len(PALETTE)])
        for k, h in enumerate(hdr, start=1):
            ws.column_dimensions[get_column_letter(k)].width = min(
                46, max(10, len(h) + 2))
        if "result" in hdr:
            ri_col = hdr.index("result") + 1
            for ri in range(2, len(rows) + 2):
                c = ws.cell(row=ri, column=ri_col)
                c.fill = PatternFill("solid", fgColor=(
                    "FFC6EFCE" if c.value == "PASS" else "FFFFC7CE"))
    buf = io.BytesIO()
    wb.save(buf)
    # Normalise the archive so the workbook is byte-reproducible: zip member
    # timestamps, and the dcterms:created / dcterms:modified stamps that
    # openpyxl rewrites to wall-clock time during save().
    buf.seek(0)
    src = zipfile.ZipFile(buf)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = re.sub(rb">\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z<",
                              b">2000-01-01T00:00:00Z<", data)
            zi = zipfile.ZipInfo(info.filename, date_time=(1980, 1, 1, 0, 0, 0))
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = info.external_attr
            dst.writestr(zi, data)
    src.close()


def build_readme(args, Y, W, weights, qcd_stream, qcd_nodes, H30, summary, SE,
                 res_parts, dropped_names, b_subset_share, optional, seeds,
                 r15_sizes, bs_classes, names, nominal, i_qq, i_cs):
    def sizes(k):
        v = [len(g["members"]) for g in res_parts[k]]
        if len(set(v)) == 1 and len(v) > 20:
            return f"{len(v)} groups, all of size {v[0]}"
        if len(v) > 20:
            return f"{len(v)} groups: {v[:4]} ... {v[-2:]}".replace("[", "").replace("]", "")
        return ", ".join(str(x) for x in v)
    defects = [a for a in ASSERTIONS if a["verdict"] == "DEFECT"]
    expfail = [a for a in ASSERTIONS
               if a["expectation"] == "MUST_NOT_HOLD" and a["result"] == "FAIL"]
    n_must = sum(1 for a in ASSERTIONS if a["expectation"] == "MUST_HOLD")
    fail_lines = "\n".join(
        f"- `{a['assertion']}` -- {a['detail']}" for a in defects) \
        or "- **none.** Every must-hold assertion passes."
    exp_lines = "\n".join(
        f"- `{a['assertion']}` -- {a['detail']}" for a in expfail) or "- none"
    core = {r["x_variable"]: r for r in SE if r["design"].startswith("core")}
    se_tbl = "\n".join(
        f"| {x} | {core[x]['x_values']} | {core[x]['Sxx']:.6g} | "
        f"{core[x]['se_factor_raw']:.6g} | "
        f"{core[x]['se_factor_range_standardised']:.6g} | "
        f"{core[x]['power_rank_within_design']} | "
        f"{core[x]['leverage_by_rung']} | "
        f"{core[x]['max_leverage_fraction']:.3f} "
        f"({core[x]['max_leverage_rung']}) |"
        for x in ("ln_R", "R_linear", "ln_N", "effective_entropy_nats",
                  "equal_spacing_rank"))
    voc_tbl = "\n".join(
        f"| {s['vocabulary']} | {s['resonant_rung']} | {s['R']} | {s['Q']} | "
        f"{s['N_classes']} | {s['entropy_nats']:.4f} | "
        f"{s['effective_class_count']:.2f} | {s['min_group_stream_share']:.3e} | "
        f"{s['max_group_stream_share']:.3e} | {s['max_over_min_ratio']:.1f} |"
        for s in summary)
    dropped = dropped_names[0] if dropped_names else "(none)"
    a8 = [a for a in ASSERTIONS if a["category"] == "item-8a"]
    a8_tbl = "\n".join(
        f"| {a['assertion'].split('[')[1].split(']')[0]} | "
        f"{a['assertion'].split(']')[1].strip()} | **{a['result']}** | "
        f"{a['detail']} |" for a in a8)
    l187 = [a for a in ASSERTIONS if a["category"] == "L187"]
    l187_tbl = "\n".join(f"| `{a['assertion']}` | **{a['result']}** | "
                         f"{a['detail']} |" for a in l187)
    return f"""# `hierarchy/` -- the label-granularity artefact set

Generated by `scripts/build_hierarchy.py`. **Nothing here is hand-edited.**
Re-run the script on the same pinned inputs and every CSV comes back
byte-identical; `06_hashes.csv` is how you check that.

    python3 scripts/build_hierarchy.py

Pinned inputs: `sophon @ 9dd6dd6` (`{args.sophon}`),
config `{YAML_REL}`. Class names come from the comment
block at YAML line {LINE_CLASS_NAMES}; the 30 reweighting groups come from the
`new_variables:` expressions at YAML lines {LINE_REWEIGHT_FIRST}-{LINE_REWEIGHT_LAST};
`class_weights` comes from lines 163-170 and sums to {W:.5f}.
See `../handoff/pins.md` for the full pin list.

---

## THE THREE THINGS I NEED YOU TO CHECK

### 1. R=29 and R=9 cannot both sit on one refinement chain. This is a proof, not a bug.

The brief asks for `161 > 45 > 29 > 15 > 9`. That chain is **impossible**, and
no choice of construction can rescue it:

* R=29 is fixed by the YAML. Its two-prong groups are
  `label_X_QQ` (indices 0-8, **9 classes**), `label_X_gg`, `label_X_ll`,
  `label_X_tauhtaul`, `label_X_tauhtauh`.
* R=9 is fixed by v3 (PLAN.md line 105): its two-prong group sizes are
  **3, 3, 4, 5**. Every group has at most 5 members.
* A coarsening of R=29 must contain `label_X_QQ` whole, i.e. must have a group
  of at least 9 members. R=9 has none. Therefore R=9 is **not** a coarsening of
  R=29, and neither is anything between them.

Concretely: R=9's two-prong split is *leading heavy flavour* --
`{{bb, bc, bq}}` / `{{cc, cs, cq}}` / `{{ss, qq, sq, gg}}` / `{{leptonic}}`
= 3 / 3 / 4 / 5. That is the only physically sensible 3/3/4/5 partition of the
two-prong block, and it necessarily cuts `label_X_QQ` into three pieces. So
R=9 *retains* quark-flavour information that R=29 has already destroyed. The
two rungs are genuinely incomparable, not merely mis-ordered.

**What I built instead** (and what you are being asked to ratify):

    R=161  >  R=45  >  R=15  >  R=9        <- PLAN.md 2.2 rule 1, the core ladder
                 >  R=29                   <- R=29 hangs off R=45 as a sibling of R=15

Both edges out of R=45 are asserted and pass. `R29 > R15` and `R29 > R9` are
asserted and **FAIL by construction**; they are in `04_refinement_assertions.csv`
with `chain_edge = "brief-requested edge"` so you can see exactly what broke.

This keeps *both* outcomes of Gate-0 decision **E2** alive at zero cost:
if you keep R=45, the ladder is 161/45/15/9 and R=29 is an optional extra arm;
if R=29 replaces R=45, the ladder becomes 161/29/... and you must then either
re-define R=9 (breaking v3 continuity) or accept that the bottom rung is
incomparable. **That is the decision. I have not made it for you.**

### 2. R=45 required dropping exactly one b-split. Ratify the criterion or override it.

R=45 is built as "R=29 with one more level of quark flavour restored, namely
b-content" -- plus the two-prong c-vs-light split that R=9 forces. Counting:
7 two-prong blocks (forced) + 24 reweighting groups in the 3/4-prong block
+ 15 available b-splits = **46**, one too many for the mandated 45.

The script therefore drops exactly one b-split, chosen by a stated, computed
criterion: **the split whose b-side carries the smallest nominal stream share**,
i.e. the least-populated and least-learnable b-resolution in the vocabulary.
That is `{dropped}`
({100*b_subset_share([s for s in optional if s[0] == dropped][0]):.4f}% of the
training stream on the b side -- the next smallest is
{100*b_subset_range(optional, dropped, b_subset_share):.4f}%). The criterion has
a unique minimum, which is asserted.

If you would rather drop a different split, or authorise **R=46** and lose the
round number, say so -- it is a one-line change to the criterion. Note the
arithmetic coincidence worth knowing: a *pure* b-split of the 29 reweighting
groups with **no** two-prong c-vs-light split gives exactly 45 groups. That is
probably where "45" came from, and it is not compatible with `45 > 15 > 9`.

### 3. Does my reading of v3's 27 / 25 / 29 / 35 / 30 match what v3 actually did?

I reverse-engineered the five 3/4-prong groups from the sizes alone, and they
came out exact and **in the order PLAN.md lists them**:

| PLAN size | my group | contents |
|---|---|---|
| 27 | `34P_4prong_hadronic` | QQQQ + QQgg + gggg |
| 25 | `34P_4prong_leptonic` | QQll + QQtauhtaul + QQtauhtauh + ggll + ggtauhtaul + ggtauhtauh |
| 29 | `34P_3prong_hadronic` | QQQ + QQg + Qgg + ggg |
| 35 | `34P_3prong_leptonic` | QQl + Qll + ggl + gll + Qtauhtaul + Qtauhtauh + gtauhtaul + gtauhtauh |
| 30 | `34P_semileptonic_nu` | QQlv + QQtaulv + QQtauhv |

Five sizes matching in the listed order is strong evidence, but it is still
inference. If v3 has an explicit map anywhere, diff it against
`02_rung_R9_groups.csv`. Same question for the two-prong 3/3/4/5.

---

## L187 -- the label-holdout arm, and why it is measure-neutral

`L187` is L188 with the two-prong `label_X_cs` (jet_label **5**) merged into
`label_X_qq` (jet_label **3**), giving 160 resonant groups + 27 QCD = **187
classes**. It is the label-holdout companion to endpoint S5.

The point of the arm is that it changes the label map and *nothing else*, and
that is a checkable claim, not an assertion of faith. Both merged classes sit
inside the **same** reweighting group `label_X_QQ` (jet_label 0-8, YAML line
{Y['rw_lines']['label_X_QQ']}). Sampling weights are defined per reweighting
group, so merging two classes inside one group cannot move a single jet between
sampling strata. The script verifies this by computing the realised 30-group
stream-share vector under both vocabularies and asserting **exact** equality on
all 30 entries:

| assertion | result | detail |
|---|---|---|
{l187_tbl}

So **L187+ sees a byte-identical jet stream to L188+**. The L187-vs-L188
contrast is a pure label-only intervention, which is exactly what a holdout arm
has to be.

Where L187 sits in the hierarchy: `R161 > L187res > R29`, both asserted and
both passing. L187 is **incomparable** with R=45, R=15 and R=9 -- the merged
pair `{{qq, cs}}` straddles their leading-heavy-flavour boundary, because `cs`
carries a c quark and `qq` does not. That is expected for a holdout arm and is
reported in `04_refinement_assertions.csv` rather than hidden.

## Correction-queue item 8a -- "no new simulation", substitutions verified

X->bs is withdrawn; blocker B10 is struck; `jetclass2_generation` is no longer
a pinned artefact. What remains is to verify that the substitutions landed.
Per-clause, PASS/FAIL, no narrative:

| clause | check | result | detail |
|---|---|---|---|
{a8_tbl}

Clause 8a.4 is an **existence and constructibility** check only. The sealed
`final-eval` split was not read and no occupancy was counted -- per-class counts
remain blocker **B4**.

## What to look at first

1. `04_refinement_assertions.csv` -- filter **`verdict = DEFECT`**. There are
   **{len(defects)}**. Do *not* filter on `result = FAIL`: most of the {len(ASSERTIONS)}
   rows are pairwise refinement questions that have no reason to be true (is
   R=9 a refinement of R=161? no, and that is not a problem), so `result` alone
   is misleading. The `expectation` column says what the study requires, and
   `verdict` is OK unless `result` disagrees with it. {n_must} rows are
   must-hold. Filter `chain_edge = "DELIVERED CHAIN EDGE"` to see the five
   edges that carry the design.
2. `03_rung_summary.csv` -- the effective class count `exp(H)` column. This is
   the number that shows how little of the nominal class count the sampler
   actually delivers.
3. `01_class_master.csv` -- scroll the `*_group_id` columns. They are colour
   filled in the workbook, so the block structure of every rung is visible in
   one screen.

## Files

| file | what it is |
|---|---|
| `00_README.md` | this file |
| `01_class_master.csv` | 188 rows, one per `jet_label`. Physical attributes, flavour flags, stream shares, and the group id + name at every rung. |
| `02_rung_<rung>_groups.csv` | one file per rung; one row per group, with members, stream-share bounds and a one-line physics rationale. |
| `03_rung_summary.csv` | one row per vocabulary: N, entropy, effective class count, share extremes, and the candidate x-variables. |
| `03b_trend_se_factors.csv` | **addition beyond the brief's file list.** Trend-contrast SE factors per candidate x-variable, for decision E4. Kept separate because it is indexed by (design, x-variable), not by rung, and jamming it into `03` would have meant a second row block in the same CSV. |
| `04_refinement_assertions.csv` | every check the script ran, including trivial ones: `expectation` (MUST_HOLD / MUST_NOT_HOLD / not_required), `result` (PASS/FAIL), `verdict` (OK / DEFECT), + detail. |
| `05_bs_mechanism.csv` | every class carrying both b and s, and where it merges away. |
| `06_hashes.csv` | SHA-256 of every file, the generator, the input YAML, the pinned commits and the released checkpoint. |
| `hierarchy_artefact.xlsx` | all of the above, one tab per file, frozen headers, colour-filled group-ID columns. |

## Rungs

| rung | groups | block | what it contracts |
|---|---|---|---|
| `R161` | 161 | 0-160 | nothing -- identity, Sophon's own resonant vocabulary |
| `L187` | 187 | 0-187 | one merge only: two-prong `cs` into `qq`. Label-holdout arm for S5. Full vocabulary, so its column in `01_class_master.csv` is populated for QCD rows too. |
| `R45` | 45 | 0-160 | quark flavour, partially: b-content resolved, c/s/light merged |
| `R29` | 29 | 0-160 | quark flavour, completely. **Exactly** the 29 `label_X_*` reweighting groups, so this is the unique rung whose classes are flat in the sampler |
| `R15` | 15 | 0-160 | + lepton flavour and tau decay mode, keeping prompt-vs-tau and gluon content |
| `R9` | 9 | 0-160 | + prompt-vs-tau and gluon content. Keeps prong count, hadronic-vs-leptonic, and (two-prong only) leading heavy flavour |
| `Q27` | 27 | 161-187 | nothing -- identity |
| `Q1` | 1 | 161-187 | all QCD heavy-flavour information |
| `L43res` | 16 | 0-160 | Sophon App. B.2 replication: 15 two-prong classes verbatim + one "other resonant" absorbing 15-160. **Not on the chain** -- its two-prong singletons are finer than R=45's two-prong groups while its resonant catch-all is coarser, so it is incomparable with every core rung. |
| `rand42_d1..d3` | 15 each | 0-160 | nothing physical. Random control, 27 QCD classes untouched. |

Group-size vectors (in group-id order, which is by smallest member index):

    R161      {sizes('R161')}
    L187res   {sizes('L187res')}
    R45       {sizes('R45')}
    R29       {sizes('R29')}
    R15       {sizes('R15')}
    R9        {sizes('R9')}
    L43res    {sizes('L43res')}
    rand42_d1 {sizes('rand42_d1')}

### rand42 construction

Three independent draws, seeds **{', '.join(str(s) for s in seeds)}**, recorded in
`02_rung_rand42_d*_groups.csv`. The permutation is *not* `random.shuffle`: it is a
sort by `sha256("rand42|draw<d>|<stratum>|seed=<s>|idx=<i>")`, so it is stable
across Python versions, platforms and forever.

Only the resonant block is randomised; all 27 QCD classes stay singletons.
Stratification is by the dataset's own prong topology, **two strata**:
`res2p` (0-14) and `res34p` (15-160). Within each stratum the block sizes are
copied from R=15: {'; '.join(f"{k} = {v}" for k, v in r15_sizes.items())}.

*Labelled decision:* "stratified within prong topology" (PLAN.md 2.2 rule 5)
admits a finer reading -- 2-prong / 3-prong / 4-prong / semileptonic. I used the
two-stratum reading because it is the dataset's own partition (`Res2P` /
`Res34P` file families) and because it preserves *less* physical meaning, which
makes rand42-vs-R15 a sharper count-versus-meaning contrast. Switching to four
strata is a one-line change to `strata` in the script.

---

## Column glossary

### `01_class_master.csv`

| column | meaning |
|---|---|
| `jet_label` | 0-187, the dataset's own index. Never re-derived. |
| `class_name` | verbatim from YAML line {LINE_CLASS_NAMES}. |
| `block` | `res2p` (0-14), `res34p` (15-160), `qcd` (161-187). |
| `topology_code` | the reweighting group's topology suffix, e.g. `QQQQ`, `QQlv`, `tauhtaul`. `QCD` for the background block. |
| `tokens` | the class name's suffix tokenised into physical objects, `+`-joined. This is what all the flags are computed from, so you can audit them. |
| `n_objects` / `n_visible_objects` | object count including / excluding neutrinos. Blank for QCD, which is not prong-labelled. |
| `has_b`, `has_c`, `has_s`, `has_g`, `has_lepton`, `has_tau`, `has_nu` | 0/1. `has_lepton` is true for e, mu **and** tau. `has_g` is always 0 for QCD -- **the QCD labels do not encode gluon content**, only matched b/c/s partons, with `label_QCD_light` meaning none. |
| `leading_heavy_flavour` | b > c > s > light > none. |
| `reweight_group_*` | which of the 30 groups this class falls in, how many classes share it, its `class_weights` value, and that value / {W:.5f}. |
| `class_stream_share_nominal` | group weight / {W:.5f} / number of classes in the group. **This assumes the classes within a reweighting group are equally occupied, which is NOT verified** -- see "Assumptions". |
| `class_stream_share_upper_bound` | group weight / {W:.5f}. **Exact and assumption-free.** A single class cannot exceed its group's whole share. |
| `class_weight_share_frac` | same value as `class_stream_share_nominal`, kept as its own column at **full precision** because it is endpoint S5's x-variable. Not rounded. |
| `class_weight_share_pct` | the same number as a percentage, for reading. Derived, never regress on it. |
| `log_class_weight_share` | `ln(class_weight_share_frac)`. **This is the S5 regressor.** |
| `log_class_weight_share_upper_bound` | `ln(class_stream_share_upper_bound)`. Exact and assumption-free: the true log share is at most this. |
| `<rung>_group_id`, `<rung>_group_name` | the group this class lands in at that rung. `n/a` where the rung partitions the other block (resonant rungs say nothing about 161-187 and vice versa). |
| `bs_relevant` | 1 if the class carries both b and s content. |
| `physics_rationale` | one line: what the class is, and why it lands where it does at the coarsest rung. |

### `02_rung_*_groups.csv`

`stream_share_nominal` uses the uniformity assumption.
`stream_share_exact_lower_bound` and `..._upper_bound` **do not**: the lower
bound sums the weights of reweighting groups wholly inside the rung group, the
upper bound also adds every reweighting group it merely touches. **Where the two
bounds coincide the number is exact**, which happens exactly when the rung group
is a union of whole reweighting groups.

Do not assume that is true of a whole rung. It holds for every group of R=29
and Q=1, and for the 3/4-prong groups of R=15 and R=9 -- but **not** for the
leading-heavy-flavour two-prong groups of R=45, R=15 and R=9, which cut inside
`label_X_QQ` and are therefore bounded, not pinned. `03_rung_summary.csv` counts
this per vocabulary in `n_groups_with_exact_share` and
`n_groups_share_bounded_only`; read those two columns before quoting any
`_nominal` number. Where the value is bounded only, the truth needs the realised
file counts (blocker **B4**).

### `03_rung_summary.csv`

`entropy_nats` is the Shannon entropy of the vocabulary's nominal group-share
prior, in nats. `effective_class_count` is `exp(H)`: the number of classes a
uniform vocabulary would need to carry the same entropy. Compare it to
`N_classes` -- the gap is how much of the nominal granularity the sampler
actually pays for.

For reference, the 30-group prior itself has **H = {H30:.6f} nats,
exp(H) = {math.exp(H30):.4f}** -- so Sophon's own sampler delivers the
diversity of about {math.exp(H30):.1f} equally-weighted groups, not 30.

| vocabulary | rung | R | Q | N | H (nats) | exp(H) | min share | max share | max/min |
|---|---|---|---|---|---|---|---|---|---|
{voc_tbl}

### `03b_trend_se_factors.csv` -- input to decision E4

For a trend contrast (OLS slope of an outcome on `x` across rungs),
`SE(slope) = sigma / sqrt(Sxx)` with `Sxx = sum (x - xbar)^2`. The **SE factor**
is `1/sqrt(Sxx)`: the multiplier on the residual SD, so **smaller is better**.

Raw SE factors are not comparable across x-variables because the units differ
(`R_linear` spans 152, `ln_R` spans 2.9). The comparable column is
`se_factor_range_standardised`, computed after mapping x to [0, 1]. That
isolates the question E4 actually asks: **which spacing makes the design
balanced?** A range-standardised `Sxx` near `n/4` is a well-spread design;
much smaller means one rung is doing all the work.

Core ladder (R = 161, 45, 15, 9 at Q = 27):

| x | values | Sxx | SE factor (raw) | SE factor (range-standardised) | power rank | leverage per rung | max leverage |
|---|---|---|---|---|---|---|---|
{se_tbl}

The five-rung extended design (with R=29) is in the same file.

**Read the leverage columns before the SE columns.** The range-standardised SE
factors span only about 8% across all five x-variables, so on precision alone
the choice barely matters. What differs sharply is *balance*: `R_linear` puts
{100*core['R_linear']['max_leverage_fraction']:.0f}% of the trend on the single
R=161 rung, so a linear-in-R trend contrast is very nearly a two-point
comparison wearing a regression's clothes. `equal_spacing_rank` is the most
balanced. `ln_R` and `ln_N` sit in between and are the defensible compromise.
That is the substance of E4, and it is a balance argument, not a power argument.

### `05_bs_mechanism.csv`

One row per class carrying both b and s, plus one `absent_from_vocabulary` row
for `label_X_bs`. `<rung>_isolated = 1` means the class is alone in its group at
that rung, i.e. b+s is still separable. `first_rung_where_bs_merges_away` names
the first rung (in the column order R161, R45, R29, R15, R9, Q27, Q1) at which
it is not.

This file is the evidence artefact for correction-queue item **8a** ("no new
simulation -- verify the substitutions landed"). It is no longer a mechanism
table for a flagship measurement; X->bs is withdrawn.

**For the record, and contrary to a claim worth killing early:** b+s content is
**not** unseen at any rung. The two-prong `label_X_bs` class genuinely does not
exist -- the two-prong list is bb, cc, ss, qq, bc, cs, bq, cq, sq, and `bs` is
the one missing pair -- but b+s appears in {len(bs_classes)} classes across the
3-prong resonant, 4-prong resonant and QCD blocks:
{', '.join(names[i] for i in bs_classes)}.

**The correct mechanism for the absence**, which any prose in the paper must
use: the neutral parent decays flavour-diagonally and the charged parent
up-type-to-down-type, so b and s can only co-occur via **two different parent
decays**, and that needs at least three prongs. Do **not** attribute it to "an
FCNC the 2HDM-like generation lacks" -- that wording appears nowhere in Sophon
App. A, whose case (1) uses `heft`.

---

## Assumptions, stated once and plainly

1. **Within-group uniformity is an assumption, not a fact.** `class_weights`
   pins the sampled fraction of each of the 30 *reweighting* groups exactly.
   It says nothing about how that fraction divides among the classes inside a
   group -- that follows the actual file occupancy, which is blocker **B4** and
   requires the corpus on the cluster PVC. Every `*_nominal` column assumes
   uniformity. Every `*_upper_bound` column does not, and is exact. Where the
   uniformity assumption is load-bearing, use the bounds.
1a. **This is a live limitation on endpoint S5, not a footnote.** S5 is the
   rarity slope: per-class P1 deficit regressed on log per-class pre-training
   stream share. Its x-variable is `log_class_weight_share`, and that column is
   **currently an upper-bound proxy, not a measurement**. `class_weights` pins
   only the 30 reweighting groups; the split *within* a group follows realised
   file occupancy. What is exact and assumption-free is
   `log_class_weight_share_upper_bound` -- the true log share is at most that
   value, for every class. The nominal column equals the upper bound minus
   `ln(n_g)`, so the proxy is a uniform downward shift **within** a reweighting
   group and a *biased* one **across** groups of different size. That matters
   for S5 specifically: a slope fitted on the proxy is contaminated by
   reweighting-group size wherever occupancy is non-uniform. Refit S5 once
   B4's realised counts exist, and until then report it as provisional. The
   ranking of classes by rarity is preserved within a group but not necessarily
   across groups.
2. **Entropy and `exp(H)` inherit that assumption**, because they are computed
   from nominal shares. The 30-group entropy `H = {H30:.6f}` nats is exact and
   assumption-free; per-rung values below R=29 are only as good as uniformity.
3. **The flat (pT, mSD) reweighting flattens *between* the 30 groups, not
   between the 188 classes.** Residual mass information survives inside a
   group. R=29 is the only rung where the label scheme and the sampler agree.
4. **QCD labels are read from `jet_label`, never re-derived.**
   `aux_genpart_isQcdParton` uses **10 GeV** (`FatJetMatching.h:591`,
   `Status == 71 && PT > 10`), the *same* threshold the labelling uses — the
   "5 GeV" once recorded here was a README error, and no 5 GeV threshold exists
   anywhere in the generation repo. **There is no differential**: label counts
   and `qcdPartons` come from the same loop. The rule stands anyway, and for a
   stronger reason: a mismatch of this kind is **structurally undetectable from
   the released data**, so re-deriving `jet_label` from auxiliary variables
   could reconstruct a different partition with nothing to warn you.
5. **QCD share, asserted:** {weights['label_QCD']} / {W:.5f} =
   **{qcd_stream*100:.4f}%** of the sampled training stream, against
   27/188 = **{qcd_nodes*100:.4f}%** of vocabulary nodes. QCD is
   {100*(qcd_stream-qcd_nodes):.4f} pp over-represented in the stream relative
   to its share of the label space.
6. **`L43res`, `L187` and `rand42_*` are deliberately off-chain.** They are
   arms, not rungs, and the assertion matrix reports their incomparability
   rather than hiding it. `L187` is the one exception that *is* comparable
   upward: `R161 > L187res > R29`.

## Defects

{fail_lines}

## Expected failures, confirmed

These are checks the study asked for and that the script confirms are
impossible. They are `MUST_NOT_HOLD`, so failing is the correct outcome.

{exp_lines}
"""


def b_subset_range(optional, dropped, fn):
    vals = sorted(fn(s) for s in optional)
    return vals[1] * 100 if len(vals) > 1 else 0.0


def build_pins(args, repo, head):
    def cmd(*a):
        return subprocess.run(a, capture_output=True, text=True).stdout.strip()

    clones = []
    for c in ("~/sophon", "~/dev/sophon", "~/e1_refs/sophon"):
        p = os.path.expanduser(c)
        if os.path.isdir(p):
            h = cmd("git", "-C", p, "rev-parse", "HEAD")
            d = cmd("git", "-C", p, "status", "--porcelain")
            tr = [l for l in d.split("\n") if l and not l.startswith("??")]
            if tr:
                state = f"**{len(tr)} TRACKED FILE(S) MODIFIED**"
            elif d:
                state = f"clean (only {len(d.splitlines())} untracked, e.g. .DS_Store)"
            else:
                state = "clean"
            yp = os.path.join(p, YAML_REL)
            yh = sha256(yp)[:12] if os.path.exists(yp) else "n/a"
            clones.append(f"| `{c}` | `{h}` | {state} | "
                          f"{'agrees' if h == SOPHON_COMMIT else '**MISMATCH**'} "
                          f"| `{yh}...` |")
    wc = os.path.expanduser("~/e1_refs/weaver-core")
    wc_head = cmd("git", "-C", wc, "rev-parse", "HEAD") if os.path.isdir(wc) else "n/a"
    wc_branch = cmd("git", "-C", wc, "rev-parse", "--abbrev-ref", "HEAD") \
        if os.path.isdir(wc) else "n/a"
    wc_ver = ""
    sp = os.path.join(wc, "setup.py")
    if os.path.exists(sp):
        m = re.search(r"version\s*=\s*'([^']+)'", open(sp).read())
        wc_ver = m.group(1) if m else ""
    ck = os.path.join(repo, "models/JetClassII_Sophon/model.pt")
    ck_hash = sha256(ck) if os.path.exists(ck) else "n/a"
    ck_size = os.path.getsize(ck) if os.path.exists(ck) else 0

    pkgs = []
    for mod in ("numpy", "pandas", "openpyxl", "yaml"):
        try:
            m = __import__(mod)
            pkgs.append(f"| `{mod}` | {getattr(m, '__version__', 'unknown')} |")
        except Exception:
            pkgs.append(f"| `{mod}` | not installed |")

    return f"""# Pins

Written by `scripts/build_hierarchy.py`. Everything below was read off this
machine at build time.

## sophon (analysis repo)

Pinned commit: **`{SOPHON_COMMIT}`** (2024-08-16).
Config used: `{YAML_REL}`.

| clone | HEAD | tree | vs pin | sha256 of the config |
|---|---|---|---|---|
{chr(10).join(clones)}

The config is byte-identical in every clone, which is what makes "the pin" mean
something. Untracked `.DS_Store` files do not affect it.

## weaver-core

**`train_sophon.sh` does not pin a commit.** It invokes the bare `weaver`
console script (`CMD="weaver"`, or `$(which weaver)` under `torchrun`) and
never names a version, tag or SHA. What it *implies* is a weaver-core that
supports these flags, all of which are non-default:

    --no-remake-weights   --data-split-num   --fetch-step
    --samples-per-epoch   --samples-per-epoch-val
    --optimizer ranger    --export-onnx      -o export_embed True
    --data-config / --network-config / --model-prefix / --backend nccl

The only version constraint in the Sophon repo is prose, in `README.md`
line 155, and it pins a **branch, not a commit**:

    pip install git+https://github.com/hqucms/weaver-core.git@dev/custom_train_eval

with the note "We are temporarily using a development branch of `weaver`."
A branch pin is not reproducible: the branch tip moves.

What is actually on this machine:

| item | value |
|---|---|
| local clone | `{wc}` |
| HEAD | `{wc_head}` |
| branch | `{wc_branch}` |
| `setup.py` version | `{wc_ver or 'unknown'}` |

Open discrepancy to resolve before launch: this clone reports
`{wc_ver or 'unknown'}`, while the project's own cluster training notes refer to
**weaver 0.4.17** quirks. The training image
`gitlab-registry.nrp-nautilus.io/escheuller/transfer-learning:cu121` is where
the runtime version actually lives, and it cannot be inspected from this
machine. **Record the resolved `weaver.__version__` and the installed commit
from inside that image, and pin it, before any production run.**
`pyproject.toml` only says `weaver-core>=0.4`, which is not a pin either.

## Released Sophon checkpoint

| item | value |
|---|---|
| source | `https://huggingface.co/jet-universe/sophon/resolve/main/models/JetClassII_Sophon/model.pt` |
| local path | `models/JetClassII_Sophon/model.pt` |
| size | {ck_size} bytes ({ck_size/1024/1024:.2f} MiB) |
| sha256 | `{ck_hash}` |

Fetched by `k8s/job-extract-embeddings.yaml` and `k8s/job-train-sweep.yaml`
via `curl` from HuggingFace `main` -- also a moving reference. The sha256 above
is the pin; check it after any re-download.

Architecture context: `train_sophon.sh` builds the model as
`networks/example_ParticleTransformer_sophon.py --use-amp -o num_classes 188
-o fc_params [(512,0.1)]`, batch 512, start-lr 5e-4, 80 epochs, optimizer
ranger, 10000*1024 samples per epoch per GPU.

## Build environment

| item | value |
|---|---|
| python | {sys.version.split()[0]} ({platform.python_implementation()}) |
| executable | `{sys.executable}` |
| platform | {platform.platform()} |

| package | version |
|---|---|
{chr(10).join(pkgs)}

The CSVs depend only on the standard library. `openpyxl` is used for the
workbook only.

## Withdrawn pins

- **`jetclass2_generation` is NOT a pinned artefact.** X->bs is withdrawn and
  blocker B10 is struck, so no new simulation is in scope and the generation
  repository is out of the provenance chain. It was never added to this file;
  this line exists so that absence is deliberate rather than an oversight.

## Not determinable locally

- **Per-class occupancy of the 188 classes (blocker B4).** Needs the Parquet
  corpus on the cluster PVC. Until it exists, per-class stream shares are
  nominal-under-uniformity with exact upper bounds; see `hierarchy/00_README.md`.
- **The runtime weaver-core commit**, as above: it lives inside the training
  image.
"""


if __name__ == "__main__":
    sys.exit(main())
